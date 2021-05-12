from openpyxl import load_workbook
import getpass

## 불러올 엑셀 파일 설정

# 현재 컴퓨터의 유저명 들고오기
userName = getpass.getuser()

# oneDrive 주소 가져오기
path1 = "C:/Users/"
path2 = "/OneDrive/AI_List.xlsx"

path = path1 + str(userName) + path2

# data_only=True : 수식없이 값만 가져오게 설정
# 내가 저장한 AI_List 들고오기
load_wb = load_workbook(path, data_only=True)

# 불러올 시트 설정
load_sheet = load_wb['매수추천']

StockList = []  # 인공지능 종목 리스트
webPageList = []  # 인공지능 추천가


# 같은 이름의 종목명 있는지 찾아서 삭제하기
def delete_stock(stock):
    for i in range(2, 100):
        # 'B2'부터 'B100'까지 읽기
        stock_name = load_sheet.cell(i, 2).value
        # 셀 비었으면 그대로 종료
        if stock_name == None:
            break
        # 종목명이 일치하면 행 삭제
        if stock_name == stock:
            load_sheet.delete_rows(i)